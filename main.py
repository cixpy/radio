
STATIONS = {
    "Band_FM":      "https://stm.alphanetdigital.com.br:7040/band",
    "Ondas_Verdes": "https://live3.livemus.com.br:6922/stream",
}

RECORD_DURATION    = 60
GROQ_WHISPER_MODEL = "whisper-large-v3-turbo"
GROQ_LLM_MODEL     = "llama-3.1-8b-instant"
MIN_SPEECH_RATIO   = 0.40
MIN_SPEECH_SEGS    = 3
TRANSCRIPTION_CAP  = 3000
AD_COOLDOWN_SECONDS = 90

# Palavras que identificam um estabelecimento varejista âncora.
# Quando um desses anunciantes é detectado, marcas mencionadas no mesmo
# bloco de áudio são tratadas como PRODUTOS do varejista, não como
# anunciantes independentes.
RETAIL_KEYWORDS = [
    "supermercado", "mercado", "atacado", "atacarejo", "hipermercado",
    "mercadão", "sacolão", "empório", "mercearia", "distribuidora",
    "magazine", "shopping", "loja", "lojas", "clube", "fair", "feira",
    "armazém", "cooperativa", "hortifruti", "quitanda",
]

import subprocess, datetime, os, re, time, shutil, json, unicodedata
import threading, queue, traceback
from zoneinfo import ZoneInfo
from dotenv import load_dotenv
import librosa, torch
from groq import Groq
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()
GROQ_API_KEY = os.getenv("GROQ_API_KEY")
if not GROQ_API_KEY:
    raise ValueError("❌ GROQ_API_KEY não encontrada!")

groq_client = Groq(api_key=GROQ_API_KEY)
TZ_BR = ZoneInfo("America/Sao_Paulo")

def br_now():       return datetime.datetime.now(TZ_BR)
def br_timestamp(): return br_now().strftime("%d-%m-%Y_%H-%M-%S")
def br_display():   return br_now().strftime("%d/%m/%Y %H:%M:%S")

def safe_filename(text: str, max_len: int = 60) -> str:
    if not text: return "Desconhecido"
    text = unicodedata.normalize("NFKD", str(text).strip())
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^\w\- ]+", "_", text, flags=re.UNICODE)
    text = re.sub(r"\s+", "_", text).strip("_")
    return (text or "Desconhecido")[:max_len]


# ─── Heurística ───────────────────────────────────────────────────────────────

# Palavras com peso 2 (forte sinal de anúncio)
AD_STRONG = [
    "promoção", "oferta", "desconto", "imperdível", "compre", "aproveite",
    "garanta", "só hoje", "últimos dias", "parcel", "sem juros", "cupom",
    "por apenas", "a partir de", "r$", "reais",
]
# Palavras com peso 1 (contexto comercial)
AD_BUSINESS = [
    "farmácia", "drogaria", "autoescola", "supermercado", "clínica",
    "laboratório", "ótica", "academia", "concessionária", "pizzaria",
    "restaurante", "seguro", "consórcio", "financiamento", "imobiliária",
    "hospital", "posto", "faculdade", "curso", "aplicativo", "frete",
    "delivery", "whatsapp", "site", ".com", ".br",
]
# CTAs que ancoram confiança
EXPLICIT_CTA = [
    "ligue", "acesse", "compre", "whatsapp", "zap", "visite", "peça já",
    "clique", "baixe", "chame no", "manda mensagem", "fale com",
    "entre em contato", "vá ao site", "pelo site", "no instagram",
    "no aplicativo", "mande um", "encomende",
]
# Frases que sinalizam conteúdo NÃO publicitário
NON_AD_PHRASES = [
    "segundo informações", "de acordo com", "o governador", "o prefeito",
    "a polícia", "os bombeiros", "o presidente", "a secretaria",
    "boa tarde", "bom dia", "boa noite", "você está ouvindo",
    "eu sinto falta", "lembro quando", "antigamente", "era melhor",
    "quando eu era", "que saudade", "minha avó", "meu pai", "minha mãe",
    "hoje em dia", "mudou muito", "apoio de", "com o apoio",
    "transmissão oficial", "patrocínio de", "parceiro oficial",
    # Auto-promoção de rádio
    "nas redes sociais", "nosso instagram", "nosso facebook",
    "siga a gente", "nos siga", "acompanhe a gente",
    "você está ouvindo a", "aqui é a", "essa é a",
    "nossa programação", "nossa rádio", "pelo nosso aplicativo",
]

PRICE_RE = re.compile(r"r\$\s*\d+([.,]\d{2})?|\d+\s*reais", re.IGNORECASE)
PHONE_RE = re.compile(r"(\(?\d{2}\)?\s*)?\d{4,5}[-\s]?\d{4}")
VINHETA_RE = re.compile(
    r"(oficial (da|do|de)|parceiro oficial|só (da|do)|"
    r"apresenta(do por)?|uma realização|com o apoio|apoio (da|do)|"
    r"patrocínio (da|do)|transmissão oficial)",
    re.IGNORECASE,
)
FIRST_PERSON_RE = re.compile(
    r"\b(eu |a gente |nós |minha |meu |nossa |nosso )"
    r"(sinto|lembro|acho|gosto|quero|fazia|comia|ia|era|fui|vim|tenho|tinha)\b",
    re.IGNORECASE,
)


def heuristic_score(text: str) -> dict:
    t = text.lower()
    return {
        "ad_score": (
            sum(2 for k in AD_STRONG    if k in t)
            + sum(1 for k in AD_BUSINESS if k in t)
            + (3 if PRICE_RE.search(t) else 0)
            + (2 if PHONE_RE.search(t) else 0)
            + (2 if any(k in t for k in EXPLICIT_CTA) else 0)
        ),
        "nonad_score": (
            sum(1 for k in NON_AD_PHRASES if k in t)
            + (3 if FIRST_PERSON_RE.search(text) else 0)
            + (4 if VINHETA_RE.search(text) and not any(k in t for k in EXPLICIT_CTA)
                 and not PRICE_RE.search(t) else 0)
        ),
        "has_price":    bool(PRICE_RE.search(t)),
        "has_phone":    bool(PHONE_RE.search(t)),
        "has_cta":      any(k in t for k in EXPLICIT_CTA),
        "is_vinheta":   bool(VINHETA_RE.search(text)),
        "is_fp_chat":   bool(FIRST_PERSON_RE.search(text)),
    }


def should_skip(heur: dict, station_name: str = "", text: str = "") -> str | None:
    """Retorna motivo de descarte ou None se deve prosseguir."""
    if heur["ad_score"] < 2 and heur["nonad_score"] >= 2:
        return "conteúdo não-publicitário"
    if heur["is_vinheta"] and not heur["has_cta"] and not heur["has_price"]:
        return "vinheta/patrocínio sem CTA ou preço"
    if heur["is_fp_chat"] and not heur["has_price"] and not heur["has_cta"]:
        return "conversa em 1ª pessoa sem preço/CTA"
    # Auto-promoção da própria rádio (menciona o nome da estação sem preço)
    if station_name and text and name_in_text(station_name, text) and not heur["has_price"]:
        return "auto-promoção da rádio"
    return None


def name_in_text(name: str, text: str) -> bool:
    """
    Verifica se o nome do anunciante aparece no texto.
    - Tokens com pelo menos 3 caracteres (era 4)
    - Exige match em pelo menos 60% dos tokens
    """
    if not name or not text:
        return False
    tokens = [w for w in re.split(r"\W+", name.lower()) if len(w) >= 3]
    t_lower = text.lower()
    if not tokens:
        return bool(re.search(r"\b" + re.escape(name.lower()) + r"\b", t_lower))
    matched = sum(1 for tok in tokens if tok in t_lower)
    return matched >= max(1, round(len(tokens) * 0.6))


def is_retail_anchor(anunciante: str, text: str) -> bool:
    """
    Retorna True se o anunciante é um estabelecimento varejista
    (supermercado, loja, clube, etc.).
    Verifica tanto o nome do anunciante quanto o contexto do texto.
    """
    if not anunciante:
        return False
    combined = (anunciante + " " + text[:500]).lower()
    return any(kw in combined for kw in RETAIL_KEYWORDS)


# ─── Worker de gravação ───────────────────────────────────────────────────────

def recorder_worker(name, url, audio_path, work_queue, stop_event):
    print(f"  🎙️  Gravador iniciado: {name}")
    while not stop_event.is_set():
        ts        = br_timestamp()
        file_path = os.path.join(audio_path, f"{safe_filename(name)}_{ts}.mp3")
        cmd = [
            "ffmpeg", "-i", url, "-t", str(RECORD_DURATION),
            "-acodec", "libmp3lame", "-ar", "16000", "-ac", "1",
            file_path, "-y", "-loglevel", "quiet",
        ]
        try:
            subprocess.run(cmd, check=True, timeout=RECORD_DURATION + 15)
            if os.path.exists(file_path):
                work_queue.put((name, file_path))
        except Exception as e:
            print(f"  ⚠️  [{name}] Erro na gravação: {e}")
            if os.path.exists(file_path):
                os.remove(file_path)
        time.sleep(2)
    print(f"  🛑 Gravador encerrado: {name}")


# ─── Helpers de estilo Excel ──────────────────────────────────────────────────

def _cell_style(c, font=None, fill=None, align=None, border=None):
    if font:   c.font      = font
    if fill:   c.fill      = fill
    if align:  c.alignment = align
    if border: c.border    = border

_THIN_BORDER = Border(**{s: Side(style="thin", color="CCCCCC")
                         for s in ("left", "right", "bottom")})
_DATA_BORDER = Border(**{s: Side(style="thin", color="DDDDDD")
                         for s in ("left", "right", "bottom")})
_DATA_FONT   = Font(name="Arial", size=10)
_CONF_FILL   = {
    "alta":  PatternFill("solid", start_color="E2EFDA"),
    "media": PatternFill("solid", start_color="FFF2CC"),
    "baixa": PatternFill("solid", start_color="FCE4D6"),
}


# ─── Classe principal ─────────────────────────────────────────────────────────

class AdDetector:
    def __init__(self):
        self.base_path   = "radio_capture"
        self.audio_path  = os.path.join(self.base_path, "temp_audios")
        self.log_path    = os.path.join(self.base_path, "logs")
        self.ads_path    = os.path.join(self.base_path, "detected_ads")
        self.report_path = os.path.join(self.base_path, "relatorio_anuncios.xlsx")

        for folder in [self.audio_path, self.log_path, self.ads_path]:
            os.makedirs(folder, exist_ok=True)

        # Cache de cooldown: chave → último datetime de detecção
        self._recent_ads: dict[str, datetime.datetime] = {}

        print("🔧 Carregando Silero VAD...")
        self.vad_model, utils = torch.hub.load(
            repo_or_dir="snakers4/silero-vad", model="silero_vad", trust_repo=True,
        )
        self.get_speech_timestamps = utils[0]
        print(f"☁️  Groq | Whisper: {GROQ_WHISPER_MODEL} | LLM: {GROQ_LLM_MODEL} | Ciclo: {RECORD_DURATION}s")
        self._init_excel()
        print("✅ Pronto.\n")

    # ── Excel ─────────────────────────────────────────────────────────────────

    def _init_excel(self):
        if os.path.exists(self.report_path):
            print(f"📊 Relatório existente: {self.report_path}"); return

        wb = Workbook()
        ws = wb.active
        ws.title = "Anúncios Detectados"
        headers = ["Data/Hora", "Rádio", "Anunciante", "Produto/Serviço",
                   "Confiança", "Trecho do Anúncio", "Arquivo de Áudio"]
        h_fill  = PatternFill("solid", start_color="1F4E79")
        h_font  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        h_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            _cell_style(c, font=h_font, fill=h_fill, align=h_align, border=_THIN_BORDER)
        ws.row_dimensions[1].height = 60
        ws.freeze_panes = "A2"
        for i, w in enumerate([20, 15, 22, 22, 12, 60, 40], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws2 = wb.create_sheet("Resumo por Rádio")
        h2_font = Font(bold=True, color="FFFFFF", name="Arial")
        h2_fill = PatternFill("solid", start_color="2E75B6")
        for col, h in enumerate(["Rádio", "Total de Anúncios", "Última Detecção"], 1):
            _cell_style(ws2.cell(row=1, column=col, value=h), font=h2_font, fill=h2_fill)

        wb.save(self.report_path)
        print(f"📊 Relatório criado: {self.report_path}")

    def _append_to_excel(self, station, info, audio_file):
        try:
            wb  = load_workbook(self.report_path)
            ws  = wb["Anúncios Detectados"]
            row = ws.max_row + 1
            conf = info.get("confianca", "baixa")
            fill = _CONF_FILL.get(conf, PatternFill("solid", start_color="FFFFFF"))
            values = [
                br_display(), station,
                info.get("anunciante") or "—",
                info.get("produto")    or "—",
                conf.upper(),
                info.get("trecho") or "—",
                os.path.basename(audio_file),
            ]
            for col, val in enumerate(values, 1):
                c = ws.cell(row=row, column=col, value=val)
                _cell_style(c, font=_DATA_FONT, fill=fill, border=_DATA_BORDER,
                            align=Alignment(vertical="center", wrap_text=(col == 6)))
            ws.row_dimensions[row].height = 18

            ws2 = wb["Resumo por Rádio"]
            sr  = next((r for r in ws2.iter_rows(min_row=2) if r[0].value == station), None)
            if sr:
                sr[1].value = (sr[1].value or 0) + 1
                sr[2].value = br_display()
            else:
                nr = ws2.max_row + 1
                for col, val in enumerate([station, 1, br_display()], 1):
                    ws2.cell(row=nr, column=col, value=val)

            wb.save(self.report_path)
            print(f"  ✅ Excel salvo ({row - 1} anúncios)")
        except Exception as e:
            print(f"  ⚠️  Erro ao salvar Excel: {e}"); traceback.print_exc()

    # ── VAD ───────────────────────────────────────────────────────────────────

    def analyze_vad(self, file_path) -> dict | None:
        try:
            y, sr = librosa.load(file_path, sr=16000)
            segs  = self.get_speech_timestamps(
                torch.from_numpy(y).float(), self.vad_model, sampling_rate=16000,
            )
            if not segs: return None
            total_speech = sum((s["end"] - s["start"]) / sr for s in segs)
            ratio = total_speech / (len(y) / sr)
            if ratio < MIN_SPEECH_RATIO or len(segs) < MIN_SPEECH_SEGS:
                return None
            return {"speech_ratio": ratio, "fragments": len(segs)}
        except Exception as e:
            print(f"  ⚠️  Erro VAD: {e}"); return None

    # ── Transcrição ───────────────────────────────────────────────────────────

    def transcribe(self, file_path) -> str:
        try:
            with open(file_path, "rb") as f:
                r = groq_client.audio.transcriptions.create(
                    file=(os.path.basename(file_path), f),
                    model=GROQ_WHISPER_MODEL, language="pt", response_format="text",
                )
            return (r if isinstance(r, str) else str(r)).strip()
        except Exception as e:
            print(f"  ⚠️  Erro Whisper: {e}"); traceback.print_exc(); return ""

    # ── Classificação ─────────────────────────────────────────────────────────

    def classify(self, text: str, heur: dict) -> list:
        """Chama o LLM e retorna lista de anúncios aprovados."""

        # Shortcut: score muito alto sem precisar do LLM
        if heur["ad_score"] >= 8 and heur["has_price"] and heur["has_cta"]:
            return [{"anunciante": None, "produto": None, "confianca": "media",
                     "trecho": "", "motivo_curto": "Score heurístico alto (sem LLM)"}]

        nivel = (
            "ATENÇÃO: alta probabilidade de anúncio real." if heur["ad_score"] >= 4
            else "Pode haver anúncio, avalie com cuidado." if heur["ad_score"] >= 2
            else "Score baixo — seja conservador."
        )
        prompt = (
            f"Classifique anúncios publicitários nesta transcrição de rádio brasileiro ({RECORD_DURATION}s).\n"
            f"Pode haver zero, um ou mais anúncios distintos.\n\n"
            f"REGRAS OBRIGATÓRIAS:\n"
            f"1. Anúncio = empresa/marca REAL com intenção de venda, promoção ou CTA.\n"
            f"2. 'anunciante' = NOME DA EMPRESA OU MARCA que paga o anúncio\n"
            f"   (ex: 'Ferreira Decorações', 'Clube Max', 'Farmácia São João').\n"
            f"   Deve aparecer LITERALMENTE no texto; caso contrário use null.\n"
            f"   NUNCA coloque descrições genéricas como 'loja' ou 'empresa'.\n"
            f"3. 'produto' = O QUE está sendo vendido/promovido\n"
            f"   (ex: 'cortinas e tapetes', 'cartão de crédito', 'cursos técnicos').\n"
            f"   NUNCA repita o nome da empresa em 'produto'.\n"
            f"4. REGRA DO VAREJISTA ÂNCORA (CRÍTICA):\n"
            f"   Se o anunciante for um supermercado, mercado, clube de compras, loja ou\n"
            f"   atacado, liste SOMENTE esse estabelecimento como anunciante — mesmo que o\n"
            f"   áudio mencione marcas de produtos (Omo, Guaraná, Nestlé etc.).\n"
            f"   Essas marcas são produtos vendidos pelo varejista, NÃO anunciantes separados.\n"
            f"   Coloque as marcas/produtos mencionados no campo 'produto' do varejista.\n"
            f"   ERRADO: [{{'anunciante':'Clube Max',...}},{{'anunciante':'Omo',...}},{{'anunciante':'Guaraná Antártica',...}}]\n"
            f"   CERTO:  [{{'anunciante':'Clube Max','produto':'bebidas, Omo, Guaraná Antártica',...}}]\n"
            f"5. 'trecho' = trecho literal do áudio que comprova o anúncio (máx 80 chars).\n"
            f"6. Ignore completamente: notícias, locução esportiva, entrevistas,\n"
            f"   conversa casual, auto-promoção da própria rádio, vinhetas sem CTA.\n"
            f"7. Confiança 'alta' SOMENTE se houver CTA explícito E (preço OU telefone).\n"
            f"8. Sem CTA E sem preço/telefone → confiança máxima = 'media'.\n"
            f"{nivel}\n\n"
            f"Exemplos corretos:\n"
            f'  {{"anunciante":"Ferreira Decorações","produto":"cortinas e tapetes","confianca":"media","trecho":"Ferreira Decorações, qualidade e estilo para sua casa"}}\n'
            f'  {{"anunciante":"Clube Max","produto":"bebidas 900ml, Omo, Guaraná Antártica","confianca":"alta","trecho":"Clube Max a partir de R$ 5,99"}}\n\n'
            f"Texto:\n\"\"\"{text}\"\"\"\n\n"
            f"Responda APENAS JSON válido, sem markdown:\n"
            f'{{"anuncios":[{{"anunciante":"nome literal ou null","produto":"o que é vendido","confianca":"alta|media|baixa","trecho":"trecho literal max80chars"}}]}}\n'
            f"Se não houver anúncio: {{\"anuncios\":[]}}"
        )

        try:
            resp = groq_client.chat.completions.create(
                model=GROQ_LLM_MODEL,
                messages=[{"role": "user", "content": prompt}],
                temperature=0, max_tokens=600,
                response_format={"type": "json_object"},
            )
            raw = resp.choices[0].message.content or ""
            try:
                data = json.loads(raw)
            except Exception:
                m = re.search(r"\{.*\}", raw, flags=re.S)
                data = json.loads(m.group(0)) if m else {}

            aprovados = []
            seen_anunciantes: set  = set()
            seen_trechos:     list = []
            has_strong_anchor = heur["has_cta"] and (heur["has_price"] or heur["has_phone"])
            has_weak_anchor   = heur["has_price"] or heur["has_phone"] or heur["has_cta"]

            for ad in (data.get("anuncios") or []):
                if not isinstance(ad, dict): continue

                # ── Anunciante: validação léxica ──
                anunc_raw = (ad.get("anunciante") or "").strip()
                anunciante = None
                if anunc_raw and anunc_raw.lower() not in ("null", "none", "desconhecido", "—", ""):
                    if name_in_text(anunc_raw, text):
                        anunciante = anunc_raw
                    else:
                        print(f"  🚫 Anunciante '{anunc_raw}' não encontrado no texto — descartado.")

                produto = (ad.get("produto") or "").strip() or None
                if produto and produto.lower() in ("null", "none", ""):
                    produto = None
                # Evita produto == anunciante (LLM às vezes espelha)
                if produto and anunciante and produto.lower() == anunciante.lower():
                    produto = None

                conf = str(ad.get("confianca", "baixa")).lower().strip()
                if conf not in ("alta", "media", "baixa"): conf = "baixa"
                trecho = (ad.get("trecho") or "")[:80]

                # ── Rebaixa confiança se faltam âncoras ──
                if conf == "alta" and not has_strong_anchor:
                    conf = "media" if has_weak_anchor else "baixa"
                    print(f"  ⬇️  Confiança rebaixada → {conf}")
                elif conf == "media" and not has_weak_anchor and heur["ad_score"] < 4:
                    conf = "baixa"
                    print("  ⬇️  Confiança rebaixada → baixa")

                # ── Filtros de aceite ──
                # Exige anunciante confirmado para qualquer confiança < alta,
                # ou para alta sem âncora forte
                if not anunciante:
                    if conf in ("baixa", "media"):
                        print(f"  ⛔ Descartado: sem anunciante confirmado (conf={conf})."); continue
                    if conf == "alta" and not has_strong_anchor:
                        print("  ⛔ Descartado: alta sem âncora e sem anunciante."); continue

                # Deduplicação por anunciante no mesmo ciclo
                chave = (anunciante or "").lower()
                if chave and chave in seen_anunciantes:
                    print(f"  🔁 Anunciante duplicado no ciclo: {chave}"); continue
                if chave:
                    seen_anunciantes.add(chave)

                # Deduplicação por trecho similar no mesmo ciclo
                if trecho:
                    if any(len(set(trecho.lower().split()) & set(t.lower().split())) > 4
                           for t in seen_trechos):
                        print(f"  🔁 Trecho duplicado: {trecho[:40]}..."); continue
                    seen_trechos.append(trecho)

                # ── Cooldown global por anunciante (anti-spam entre ciclos) ──
                chave_tempo = (anunciante or trecho[:20] or "unknown").lower()
                ultimo = self._recent_ads.get(chave_tempo)
                if ultimo and (br_now() - ultimo).total_seconds() < AD_COOLDOWN_SECONDS:
                    print(f"  🕐 Cooldown ativo para '{chave_tempo}' — ignorando.")
                    continue
                self._recent_ads[chave_tempo] = br_now()

                # Promove baixa→media se score alto com anunciante confirmado
                if conf == "baixa" and heur["ad_score"] >= 6 and anunciante:
                    conf = "media"

                aceito = (
                    conf == "alta"
                    or (conf == "media" and heur["ad_score"] >= 3)
                )
                if aceito:
                    aprovados.append({"anunciante": anunciante, "produto": produto,
                                      "confianca": conf, "trecho": trecho})

            # ── Consolidação de varejista âncora ──────────────────────────────
            # Se um dos aprovados é um varejista (supermercado, loja, clube…),
            # as demais marcas do mesmo bloco são produtos dele, não anunciantes.
            retail_ads  = [a for a in aprovados if is_retail_anchor(a.get("anunciante", ""), text)]
            nonretail   = [a for a in aprovados if not is_retail_anchor(a.get("anunciante", ""), text)]
            if retail_ads and nonretail:
                # Agrupa marcas avulsas como produtos do varejista
                extra = ", ".join(
                    filter(None, [b.get("anunciante") or b.get("produto") for b in nonretail])
                )
                for ra in retail_ads:
                    base = ra.get("produto") or ""
                    ra["produto"] = (base + (", " + extra if extra else "")).strip(", ")
                print(f"  🏪 Varejista âncora: {len(nonretail)} marca(s) incorporada(s) como produto.")
                aprovados = retail_ads

            # Fallback heurístico se LLM não detectou nada mas o score é alto
            if not aprovados and heur["ad_score"] >= 7 and heur["has_price"] and heur["has_cta"]:
                aprovados.append({"anunciante": None, "produto": None, "confianca": "baixa",
                                  "trecho": "", "motivo_curto": "Detectado por heurística"})
            return aprovados

        except Exception as e:
            print(f"  ⚠️  Erro LLM: {e}"); return []

    # ── Salvar áudio ──────────────────────────────────────────────────────────

    def save_ad(self, station, audio_file, info, index=0) -> str:
        parts = [safe_filename(station), safe_filename(info.get("anunciante") or "Desconhecido")]
        produto = safe_filename(info.get("produto") or "")
        if produto and produto != "Desconhecido":
            parts.append(produto)
        parts.append(br_timestamp())
        if index > 0: parts.append(f"ad{index}")
        dest = os.path.join(self.ads_path, "__".join(parts) + ".mp3")
        shutil.copy2(audio_file, dest)
        return dest

    # ── Processar item ────────────────────────────────────────────────────────

    def process_item(self, name, audio_file):
        try:
            vad = self.analyze_vad(audio_file)
            if not vad:
                print(f"  🎵 [{name}] Ignorado (pouca fala)"); return

            print(f"  🔍 [{name}] Transcrevendo... "
                  f"(speech={vad['speech_ratio']:.0%}, frags={vad['fragments']})")

            full_text = self.transcribe(audio_file)
            snippet   = full_text[:TRANSCRIPTION_CAP]

            if not snippet:
                print(f"  ⚠️  [{name}] Transcrição vazia."); return

            heur = heuristic_score(snippet)
            print(f"  📊 [{name}] ad={heur['ad_score']} nonad={heur['nonad_score']} "
                  f"cta={heur['has_cta']} preço={heur['has_price']} vinheta={heur['is_vinheta']}")

            # Passa station_name e text para filtrar auto-promoção
            motivo = should_skip(heur, station_name=name, text=snippet)
            if motivo:
                print(f"  🎵 [{name}] Descartado: {motivo}."); return

            print(f"  📝 [{name}] {snippet[:200].replace(chr(10), ' ')!r}")
            anuncios = self.classify(snippet, heur)

            if not anuncios:
                print(f"  🎵 [{name}] Nenhum anúncio detectado."); return

            print(f"  📢 [{name}] {len(anuncios)} anúncio(s)!")
            for i, info in enumerate(anuncios, 1):
                marca = info.get("anunciante") or "Desconhecido"
                conf  = info.get("confianca", "media")
                print(f"       [{i}] {marca} (conf={conf}) — {info.get('trecho', '')[:80]}")
                saved = self.save_ad(name, audio_file, info, index=i if len(anuncios) > 1 else 0)
                self._append_to_excel(name, info, saved)
                print(f"       💾 {os.path.basename(saved)}")

        except Exception as e:
            print(f"  ❌ [{name}] Erro: {e}"); traceback.print_exc()
        finally:
            if os.path.exists(audio_file):
                os.remove(audio_file)

    # ── Loop principal ────────────────────────────────────────────────────────

    def run(self):
        print("🚀 Monitoramento iniciado...")
        print(f"   Rádios   : {', '.join(STATIONS.keys())}")
        print(f"   Duração  : {RECORD_DURATION}s | Relatório: {os.path.abspath(self.report_path)}")
        print("   (Ctrl+C para parar)\n")

        work_queue = queue.Queue()
        stop_event = threading.Event()
        threads    = []

        for name, url in STATIONS.items():
            t = threading.Thread(
                target=recorder_worker,
                args=(name, url, self.audio_path, work_queue, stop_event),
                daemon=True, name=f"rec-{name}",
            )
            t.start(); threads.append(t)

        print(f"  🎙️  {len(threads)} gravadores iniciados.\n")
        try:
            while True:
                try:
                    name, audio_file = work_queue.get(timeout=2)
                    print(f"\n{'─'*60}\n📥 [{name}] Novo áudio — {br_display()}")
                    self.process_item(name, audio_file)
                    work_queue.task_done()
                except queue.Empty:
                    continue
        except KeyboardInterrupt:
            print("\n\n🛑 Encerrando...")
            stop_event.set()
            for t in threads: t.join(timeout=5)
            print("👋 Encerrado.")


if __name__ == "__main__":
    AdDetector().run()